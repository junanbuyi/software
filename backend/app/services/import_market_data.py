"""
导入电力市场平台数据到数据库
数据来源：电力市场平台数据/ 目录下的三个Excel文件
"""
from __future__ import annotations

import json
import sys
import os

import openpyxl
from sqlalchemy.orm import Session

# 确保项目根目录在 sys.path
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")))

from app.db.session import SessionLocal, get_engine, Base
from app.models.market import (
    MarketThermalPlant,
    MarketThermalUnit,
    MarketWindUnit,
    MarketSolarUnit,
    MarketClearingHistory,
    MarketOutResult,
    MarketDayAheadQuote,
    MarketAuxiliaryQuote,
    MarketLoad,
)

DATA_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", "..", "电力市场平台数据"))

OUT_FILE = os.path.join(DATA_DIR, "out.xlsx")
INPUT_FILE = os.path.join(DATA_DIR, "ROTS 44_DM&ASInputData.xlsm")
HISTORY_FILE = os.path.join(DATA_DIR, "ROTS_合并数据_按顺序.xlsx")


def create_tables():
    """创建所有市场相关表"""
    engine = get_engine()
    Base.metadata.create_all(bind=engine)
    print("[OK] 数据库表已创建/确认存在")


def import_thermal_plants(db: Session, wb):
    """导入火电厂信息"""
    ws = wb["UnitThermalPlants"]
    count = 0
    for r in range(4, ws.max_row + 1):
        plant_id = ws.cell(r, 1).value
        name = ws.cell(r, 2).value
        if not plant_id:
            continue
        obj = MarketThermalPlant(plant_id=str(plant_id), name=str(name))
        db.merge(obj)
        count += 1
    db.commit()
    print(f"[OK] 火电厂信息: {count} 条")


def import_thermal_units(db: Session, wb):
    """导入火电机组参数"""
    ws = wb["UnitThermalGenerators"]
    count = 0
    for r in range(4, ws.max_row + 1):
        uid = ws.cell(r, 1).value
        if not uid:
            continue
        obj = MarketThermalUnit(
            unit_id=str(uid),
            unit_name=str(ws.cell(r, 2).value or ""),
            plant_id=str(ws.cell(r, 3).value or ""),
            bus_id=str(ws.cell(r, 4).value or ""),
            capacity=float(ws.cell(r, 5).value or 0),
            initial_state=bool(ws.cell(r, 6).value),
            initial_duration=int(ws.cell(r, 7).value or 0),
            min_output=float(ws.cell(r, 9).value or 0),
            ramp_up=float(ws.cell(r, 11).value or 0),
            ramp_down=float(ws.cell(r, 12).value or 0),
            min_on_time=int(ws.cell(r, 13).value or 0),
            min_off_time=int(ws.cell(r, 14).value or 0),
            start_cost=float(ws.cell(r, 15).value or 0),
            stop_cost=float(ws.cell(r, 16).value or 0),
            fuel_id=str(ws.cell(r, 17).value or ""),
            cost_a=float(ws.cell(r, 18).value or 0),
            cost_b=float(ws.cell(r, 19).value or 0),
            cost_c=float(ws.cell(r, 20).value or 0),
            freq_response=float(ws.cell(r, 21).value or 0),
            freq_error=float(ws.cell(r, 22).value or 0),
            reg_speed=float(ws.cell(r, 23).value or 0),
        )
        db.merge(obj)
        count += 1
    db.commit()
    print(f"[OK] 火电机组参数: {count} 条")


def import_wind_units(db: Session, wb):
    """导入风电机组"""
    ws = wb["UnitWindGenerators"]
    count = 0
    for r in range(4, ws.max_row + 1):
        uid = ws.cell(r, 1).value
        if not uid:
            continue
        obj = MarketWindUnit(
            unit_id=str(uid),
            unit_name=str(ws.cell(r, 2).value or ""),
            bus_id=str(ws.cell(r, 3).value or ""),
            capacity=float(ws.cell(r, 4).value or 0),
            curve_name=str(ws.cell(r, 5).value or ""),
        )
        db.merge(obj)
        count += 1
    db.commit()
    print(f"[OK] 风电机组: {count} 条")


def import_solar_units(db: Session, wb):
    """导入光伏机组"""
    ws = wb["UnitSolarGenerators"]
    count = 0
    for r in range(4, ws.max_row + 1):
        uid = ws.cell(r, 1).value
        if not uid:
            continue
        obj = MarketSolarUnit(
            unit_id=str(uid),
            unit_name=str(ws.cell(r, 2).value or ""),
            bus_id=str(ws.cell(r, 3).value or ""),
            capacity=float(ws.cell(r, 4).value or 0),
            curve_name=str(ws.cell(r, 5).value or ""),
        )
        db.merge(obj)
        count += 1
    db.commit()
    print(f"[OK] 光伏机组: {count} 条")


def import_loads(db: Session, wb):
    """导入负荷数据"""
    ws = wb["Loads"]
    count = 0
    for r in range(4, ws.max_row + 1):
        lid = ws.cell(r, 1).value
        if not lid:
            continue
        obj = MarketLoad(
            load_id=str(lid),
            load_name=str(ws.cell(r, 2).value or ""),
            bus_id=str(ws.cell(r, 3).value or ""),
            load_curve_name=str(ws.cell(r, 4).value or ""),
            capacity_ratio=float(ws.cell(r, 5).value or 0),
            responsiveness=float(ws.cell(r, 6).value or 0),
            is_interrupt=bool(ws.cell(r, 7).value),
            is_transfer=bool(ws.cell(r, 8).value),
            interrupt_cost=float(ws.cell(r, 9).value or 0),
            transfer_cost=float(ws.cell(r, 10).value or 0),
        )
        db.merge(obj)
        count += 1
    db.commit()
    print(f"[OK] 负荷数据: {count} 条")


def import_day_ahead_quotes(db: Session, wb):
    """导入日前市场报价"""
    ws = wb["MarDayAheadUnitQuotes"]
    # 先清空旧数据
    db.query(MarketDayAheadQuote).delete()
    db.commit()
    count = 0
    for r in range(4, ws.max_row + 1):
        qid = ws.cell(r, 1).value
        if qid is None:
            continue
        obj = MarketDayAheadQuote(
            quote_id=int(qid),
            unit_id=str(ws.cell(r, 2).value or ""),
            market_name=str(ws.cell(r, 3).value or ""),
            quote_time=int(ws.cell(r, 4).value or 0),
            quote_section=str(ws.cell(r, 5).value or ""),
            quote_price=float(ws.cell(r, 6).value or 0),
            quote_quantity=float(ws.cell(r, 7).value or 0),
            is_used=bool(ws.cell(r, 8).value),
        )
        db.add(obj)
        count += 1
    db.commit()
    print(f"[OK] 日前市场报价: {count} 条")


def import_auxiliary_quotes(db: Session, wb):
    """导入辅助服务报价"""
    ws = wb["MarAucillaryUnitQuotes"]
    db.query(MarketAuxiliaryQuote).delete()
    db.commit()
    count = 0
    for r in range(4, ws.max_row + 1):
        qid = ws.cell(r, 1).value
        if qid is None:
            continue
        obj = MarketAuxiliaryQuote(
            quote_id=int(qid),
            unit_id=str(ws.cell(r, 2).value or ""),
            market_name=str(ws.cell(r, 3).value or ""),
            quote_type=int(ws.cell(r, 4).value or 0),
            quote_time=int(ws.cell(r, 5).value or 0),
            quote_section=str(ws.cell(r, 6).value or ""),
            quote_price=float(ws.cell(r, 7).value or 0),
            quote_quantity=float(ws.cell(r, 8).value or 0),
            is_used=bool(ws.cell(r, 9).value),
        )
        db.add(obj)
        count += 1
    db.commit()
    print(f"[OK] 辅助服务报价: {count} 条")


def import_clearing_history(db: Session):
    """导入成交历史数据 (ROTS_合并数据_按顺序.xlsx)"""
    wb = openpyxl.load_workbook(HISTORY_FILE, data_only=True)
    ws = wb["Sheet1"]
    db.query(MarketClearingHistory).delete()
    db.commit()
    count = 0
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, 1).value
        if not name:
            continue
        values = []
        for c in range(2, ws.max_column + 1):
            v = ws.cell(r, c).value
            values.append(float(v) if v is not None else 0.0)
        obj = MarketClearingHistory(
            metric_name=str(name),
            period_values=values,
            total_periods=len(values),
        )
        db.add(obj)
        count += 1
    db.commit()
    wb.close()
    print(f"[OK] 成交历史数据: {count} 条 (每条含 {ws.max_column - 1} 个时段)")


def import_out_results(db: Session):
    """导入市场出清结果 (out.xlsx)"""
    wb = openpyxl.load_workbook(OUT_FILE, data_only=True)
    db.query(MarketOutResult).delete()
    db.commit()
    total = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        count = 0
        for r in range(2, ws.max_row + 1):
            row_idx = ws.cell(r, 1).value
            if row_idx is None:
                continue
            values = []
            for c in range(2, ws.max_column + 1):
                v = ws.cell(r, c).value
                values.append(float(v) if v is not None else 0.0)
            obj = MarketOutResult(
                sheet_name=sheet_name,
                row_index=int(row_idx),
                period_values=values,
            )
            db.add(obj)
            count += 1
        db.commit()
        total += count
        print(f"  - {sheet_name}: {count} 行")
    wb.close()
    print(f"[OK] 出清结果合计: {total} 条")


def main():
    print("=" * 50)
    print("电力市场平台数据导入")
    print("=" * 50)

    # 1. 建表
    create_tables()

    db = SessionLocal()
    try:
        # 2. 导入输入数据 (ROTS 44_DM&ASInputData.xlsm)
        print("\n--- 导入输入数据 ---")
        input_wb = openpyxl.load_workbook(INPUT_FILE, data_only=True)
        import_thermal_plants(db, input_wb)
        import_thermal_units(db, input_wb)
        import_wind_units(db, input_wb)
        import_solar_units(db, input_wb)
        import_loads(db, input_wb)
        import_day_ahead_quotes(db, input_wb)
        import_auxiliary_quotes(db, input_wb)
        input_wb.close()

        # 3. 导入成交历史 (ROTS_合并数据_按顺序.xlsx)
        print("\n--- 导入成交历史 ---")
        import_clearing_history(db)

        # 4. 导入出清结果 (out.xlsx)
        print("\n--- 导入出清结果 ---")
        import_out_results(db)

        print("\n" + "=" * 50)
        print("全部导入完成!")
        print("=" * 50)
    finally:
        db.close()


if __name__ == "__main__":
    main()
