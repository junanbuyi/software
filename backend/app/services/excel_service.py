"""Excel和CSV文件解析服务"""
from __future__ import annotations

import csv
from datetime import datetime
from io import BytesIO, StringIO
from typing import List, Dict, Any

import openpyxl


def parse_excel_file(file_content: bytes) -> List[Dict[str, Any]]:
    """
    解析Excel文件内容。
    
    期望的Excel格式：
    - 第一行为表头
    - 列：时间, 电价(元/kWh), 发电量预测(kWh), 负荷预测(kW), 天气类型, 是否节假日
    
    Args:
        file_content: Excel文件的字节内容
    
    Returns:
        解析后的记录列表
    """
    workbook = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
    sheet = workbook.active
    
    records = []
    headers = []
    
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
        if row_idx == 0:
            # 第一行是表头
            headers = [str(cell).strip() if cell else "" for cell in row]
            continue
        
        if not any(row):
            # 跳过空行
            continue
        
        record = {}
        for col_idx, cell in enumerate(row):
            if col_idx < len(headers):
                header = headers[col_idx].lower()
                
                # 映射列名到字段名
                if "时间" in header or "time" in header or "date" in header:
                    if isinstance(cell, datetime):
                        record["record_time"] = cell
                    elif cell:
                        try:
                            record["record_time"] = datetime.strptime(str(cell), "%Y-%m-%d %H:%M:%S")
                        except ValueError:
                            try:
                                record["record_time"] = datetime.strptime(str(cell), "%Y-%m-%d")
                            except ValueError:
                                record["record_time"] = None
                
                elif "电价" in header or "price" in header:
                    record["price_kwh"] = float(cell) if cell else 0.0
                
                elif "发电" in header or "generation" in header:
                    record["generation_kwh"] = float(cell) if cell else 0.0
                
                elif "负荷" in header or "load" in header:
                    record["load_kw"] = float(cell) if cell else 0.0
                
                elif "天气" in header or "weather" in header:
                    record["weather_type"] = str(cell) if cell else "unknown"
                
                elif "节假日" in header or "holiday" in header:
                    if isinstance(cell, bool):
                        record["is_holiday"] = cell
                    elif cell:
                        record["is_holiday"] = str(cell).lower() in ("是", "yes", "true", "1")
                    else:
                        record["is_holiday"] = False
        
        # 只添加有效记录
        if record.get("record_time"):
            records.append(record)
    
    return records


def parse_csv_file(file_content: bytes) -> List[Dict[str, Any]]:
    """
    解析CSV文件内容。
    
    Args:
        file_content: CSV文件的字节内容
    
    Returns:
        解析后的记录列表
    """
    # 尝试不同编码解析CSV
    text = None
    for encoding in ['utf-8', 'gbk', 'gb2312', 'utf-8-sig']:
        try:
            text = file_content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    
    if text is None:
        raise ValueError("无法解码CSV文件，请使用UTF-8或GBK编码")
    
    reader = csv.reader(StringIO(text))
    rows = list(reader)
    
    if len(rows) < 2:
        raise ValueError("CSV文件至少需要包含表头和一行数据")
    
    headers = [h.strip().lower() for h in rows[0]]
    records = []
    
    for row in rows[1:]:
        if not any(row):
            continue
        
        record = {}
        for col_idx, cell in enumerate(row):
            if col_idx >= len(headers):
                continue
            
            header = headers[col_idx]
            cell = cell.strip() if cell else ""
            
            if "时间" in header or "time" in header or "date" in header:
                if cell:
                    formats = [
                        "%Y-%m-%d %H:%M:%S",
                        "%Y-%m-%d %H:%M",
                        "%Y-%m-%d",
                        "%Y/%m/%d %H:%M:%S",
                        "%Y/%m/%d %H:%M",
                        "%Y/%m/%d",
                    ]
                    record["record_time"] = None
                    for fmt in formats:
                        try:
                            record["record_time"] = datetime.strptime(cell, fmt)
                            break
                        except ValueError:
                            continue
            
            elif "电价" in header or "price" in header:
                record["price_kwh"] = float(cell) if cell else 0.0
            
            elif "发电" in header or "generation" in header:
                record["generation_kwh"] = float(cell) if cell else 0.0
            
            elif "负荷" in header or "load" in header:
                record["load_kw"] = float(cell) if cell else 0.0
            
            elif "天气" in header or "weather" in header:
                record["weather_type"] = cell if cell else "unknown"
            
            elif "节假日" in header or "holiday" in header:
                record["is_holiday"] = cell.lower() in ("是", "yes", "true", "1") if cell else False
        
        if record.get("record_time"):
            records.append(record)
    
    return records


def validate_excel_structure(file_content: bytes) -> tuple[bool, str]:
    """
    验证Excel文件结构是否符合要求。
    
    Args:
        file_content: Excel文件的字节内容
    
    Returns:
        (是否有效, 错误信息)
    """
    try:
        workbook = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
        sheet = workbook.active
        
        # 检查是否有数据
        if sheet.max_row < 2:
            return False, "Excel文件至少需要包含表头和一行数据"
        
        # 检查表头
        headers = [str(cell).strip().lower() if cell else "" for cell in next(sheet.iter_rows(values_only=True))]
        
        required_fields = ["时间", "电价"]
        found_fields = []
        
        for header in headers:
            if "时间" in header or "time" in header or "date" in header:
                found_fields.append("时间")
            elif "电价" in header or "price" in header:
                found_fields.append("电价")
        
        missing = set(required_fields) - set(found_fields)
        if missing:
            return False, f"缺少必要的列: {', '.join(missing)}"
        
        return True, ""
    
    except Exception as e:
        return False, f"解析Excel文件失败: {str(e)}"
